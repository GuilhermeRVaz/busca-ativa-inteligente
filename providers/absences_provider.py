from pathlib import Path
from typing import Any


class ConsolidatedAbsencesProvider:
    def fetch_absences(self, report_path: str | Path, day: int) -> list[dict[str, str]]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ImportError(
                "Dependencia ausente: openpyxl. Instale com: python -m pip install -r requirements.txt"
            ) from exc

        path = Path(report_path)
        if not path.exists():
            raise FileNotFoundError(self._build_missing_report_message(path))

        workbook = load_workbook(path, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        header_index = self._find_header_index(rows)
        headers = [self._safe_text(value) for value in rows[header_index]]
        day_column_index = self._find_day_column_index(headers, day)
        class_column_index = self._find_class_column_index(headers)
        name_column_index = self._find_required_column_index(headers, {"nome"})
        ra_column_index = self._find_required_column_index(headers, {"ra"})

        absences: list[dict[str, str]] = []
        for row in rows[header_index + 1 :]:
            student_name = self._safe_text(self._get_cell(row, name_column_index))
            ra_value = self._safe_text(self._get_cell(row, ra_column_index))
            if not student_name or not ra_value:
                continue

            absence_count = self._absence_cell_to_int(self._get_cell(row, day_column_index))
            if absence_count <= 0:
                continue

            class_name = self._safe_text(self._get_cell(row, class_column_index))
            absences.append(
                {
                    "student_name": student_name,
                    "class_name": class_name,
                    "absence_days": str(day),
                }
            )

        return absences

    def _find_header_index(self, rows: list[tuple[Any, ...]]) -> int:
        for index, row in enumerate(rows):
            normalized = {self._normalize_header(cell) for cell in row if self._safe_text(cell)}
            if "nome" in normalized and "ra" in normalized:
                return index
        raise ValueError("Cabecalho do relatorio consolidado nao encontrado.")

    def _find_day_column_index(self, headers: list[str], day: int) -> int:
        day_text = str(day)
        for index, header in enumerate(headers):
            if self._normalize_header(header) == day_text:
                return index
        raise ValueError(f"Dia {day} nao encontrado no relatorio consolidado.")

    def _find_class_column_index(self, headers: list[str]) -> int:
        for index, header in enumerate(headers):
            if self._normalize_header(header) == "turma":
                return index
        return 0

    def _find_required_column_index(self, headers: list[str], names: set[str]) -> int:
        for index, header in enumerate(headers):
            if self._normalize_header(header) in names:
                return index
        raise ValueError(f"Colunas obrigatorias nao encontradas: {', '.join(sorted(names))}")

    @staticmethod
    def _get_cell(row: tuple[Any, ...], index: int) -> Any:
        if index >= len(row):
            return None
        return row[index]

    @staticmethod
    def _absence_cell_to_int(value: Any) -> int:
        if value in (None, ""):
            return 0
        try:
            return int(float(str(value).strip()))
        except (TypeError, ValueError):
            return 0

    @staticmethod
    def _safe_text(value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip()

    @staticmethod
    def _normalize_header(value: Any) -> str:
        text = str(value or "").strip().lower()
        return " ".join(text.split())

    def _build_missing_report_message(self, missing_path: Path) -> str:
        message = f"Relatorio consolidado nao encontrado: {missing_path}"
        suggestions = self._find_report_candidates(missing_path.name)
        if not suggestions:
            return (
                f"{message}. Gere o consolidado no legado primeiro ou informe "
                "--report-path com o caminho correto."
            )

        suggested_path = suggestions[0]
        return (
            f"{message}. Encontrei um consolidado em: {suggested_path}. "
            f"Teste com --report-path \"{suggested_path}\" ou ajuste CONSOLIDATED_REPORT_PATH."
        )

    @staticmethod
    def _find_report_candidates(filename: str) -> list[Path]:
        base_dir = Path.cwd().parent
        candidates = sorted(base_dir.glob(f"*/relatorios/{filename}"))
        return [candidate.resolve() for candidate in candidates if candidate.exists()]
