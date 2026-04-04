import json
import unicodedata
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

import gspread
from openpyxl import load_workbook

from core.config import settings
from core.logging import get_logger


logger = get_logger(__name__)


class LocalRepository:
    def __init__(self) -> None:
        self.base_path = Path(settings.data_dir)
        self.base_path.mkdir(parents=True, exist_ok=True)
        self.messages_file = self.base_path / "messages.json"
        self.incoming_messages_file = self.base_path / "incoming_messages.json"
        self.campaigns_file = self.base_path / "campaigns.json"
        self._contacts_cache: list[dict[str, str]] | None = None
        self._contacts_cache_key: tuple[str, str] | None = None
        self._consolidated_cache: dict[str, dict[str, str]] | None = None
        self._consolidated_cache_path: str | None = None

    def _read_json(self, file_path: Path) -> list[dict[str, Any]]:
        if not file_path.exists():
            return []
        try:
            return json.loads(file_path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            logger.warning("Invalid JSON found in %s. Resetting file.", file_path)
            return []

    def _write_json(self, file_path: Path, data: list[dict[str, Any]]) -> None:
        file_path.write_text(
            json.dumps(data, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    def _get_gspread_client(self) -> gspread.Client | None:
        credentials_file = Path(settings.google_service_account_file)
        if not credentials_file.exists():
            logger.warning(
                "Arquivo de service account nao encontrado em %s",
                credentials_file,
            )
            return None

        try:
            return gspread.service_account(filename=str(credentials_file))
        except Exception as exc:
            logger.error("Falha ao autenticar no Google Sheets: %s", exc)
            return None

    def _get_worksheets(self, spreadsheet_url: str, worksheet_setting: str) -> list[Any]:
        client = self._get_gspread_client()
        if client is None:
            return []

        try:
            spreadsheet = client.open_by_url(spreadsheet_url)
            resolved_setting = worksheet_setting.strip()
            if not resolved_setting or resolved_setting == "*":
                return list(spreadsheet.worksheets())

            worksheet_names = [
                item.strip() for item in resolved_setting.split(",") if item.strip()
            ]
            return [spreadsheet.worksheet(name) for name in worksheet_names]
        except Exception as exc:
            logger.error("Falha ao abrir planilha no Google Sheets: %s", exc)
            return []

    def save_message(
        self,
        conversation_id: str,
        direction: str,
        text: str,
        metadata: dict[str, Any] | None = None,
    ) -> None:
        messages = self._read_json(self.messages_file)
        messages.append(
            {
                "conversation_id": conversation_id,
                "direction": direction,
                "text": text,
                "metadata": metadata or {},
            }
        )
        self._write_json(self.messages_file, messages)

    def carregar_contatos(self) -> list[dict[str, str]]:
        if not settings.google_sheet_contatos_url:
            raise ValueError("Defina GOOGLE_SHEET_CONTATOS_URL para carregar contatos.")

        cache_key = (
            settings.google_sheet_contatos_url,
            settings.google_sheet_contatos_worksheet,
        )
        if self._contacts_cache is not None and self._contacts_cache_key == cache_key:
            return list(self._contacts_cache)

        worksheets = self._get_worksheets(
            settings.google_sheet_contatos_url,
            settings.google_sheet_contatos_worksheet,
        )
        if not worksheets:
            return []

        records: list[dict[str, Any]] = []
        for worksheet in worksheets:
            records.extend(worksheet.get_all_records())
        contacts = self._records_to_contacts(records)
        self._contacts_cache = contacts
        self._contacts_cache_key = cache_key
        return list(contacts)

    def salvar_interacao(self, data: dict[str, Any]) -> None:
        entry = {
            "data_hora": str(data.get("data_hora", "")).strip(),
            "student_name": str(data.get("student_name", "")).strip(),
            "class_name": str(data.get("class_name", "")).strip(),
            "ra": str(data.get("ra", "")).strip(),
            "tipo_responsavel": str(data.get("tipo_responsavel", "")).strip(),
            "numero_chamado": str(
                data.get("numero_chamado", data.get("telefone", ""))
            ).strip(),
            "identificador_remetente": str(
                data.get("identificador_remetente", data.get("telefone", ""))
            ).strip(),
            "telefone": str(data.get("telefone", "")).strip(),
            "mensagem": str(data.get("mensagem", "")),
            "classificacao": str(data.get("classificacao", "")).strip(),
            "intencao": str(data.get("intencao", "")).strip(),
            "motivo": str(data.get("motivo", "")).strip(),
            "observacao": str(data.get("observacao", "")).strip(),
            "campaign_id": str(data.get("campaign_id", "")).strip(),
            "origem": str(data.get("origem", "whatsapp")).strip() or "whatsapp",
            "raw_payload": data.get("raw_payload") or {},
        }
        if not entry["intencao"] and entry["classificacao"]:
            entry["intencao"] = entry["classificacao"]
        if not entry["intencao"]:
            entry["intencao"] = "DUVIDA"
        if not entry["motivo"]:
            entry["motivo"] = "OUTRO"
        if not entry["observacao"]:
            entry["observacao"] = "nao foi possivel classificar"
        entry = self._enrich_interaction_entry(entry)

        self._save_incoming_json(entry)
        self._save_incoming_google_sheet(entry)

    def save_incoming_message(
        self,
        telefone: str,
        mensagem: str,
        classificacao: str,
        data_hora: str,
        raw_payload: dict[str, Any] | None = None,
        campaign_id: str = "",
        origem: str = "whatsapp",
        intencao: str = "",
        motivo: str = "",
        observacao: str = "",
        student_name: str = "",
        class_name: str = "",
        ra: str = "",
        tipo_responsavel: str = "",
    ) -> None:
        self.salvar_interacao(
            {
                "telefone": telefone,
                "mensagem": mensagem,
                "classificacao": classificacao,
                "intencao": intencao,
                "motivo": motivo,
                "observacao": observacao,
                "data_hora": data_hora,
                "campaign_id": campaign_id,
                "origem": origem,
                "student_name": student_name,
                "class_name": class_name,
                "ra": ra,
                "tipo_responsavel": tipo_responsavel,
                "raw_payload": raw_payload or {},
            }
        )

    def save_campaign_event(
        self,
        campaign_id: str,
        event_type: str,
        payload: dict[str, Any] | None = None,
    ) -> None:
        campaigns = self._read_json(self.campaigns_file)
        campaigns.append(
            {
                "campaign_id": campaign_id,
                "event_type": event_type,
                "payload": payload or {},
            }
        )
        self._write_json(self.campaigns_file, campaigns)

    def _save_incoming_json(self, entry: dict[str, Any]) -> None:
        messages = self._read_json(self.incoming_messages_file)
        messages.append(entry)
        self._write_json(self.incoming_messages_file, messages)
        logger.info("backup json salvo | arquivo=%s", self.incoming_messages_file)

    def _save_incoming_google_sheet(self, entry: dict[str, Any]) -> None:
        if not settings.google_sheet_dados_url:
            logger.warning("GOOGLE_SHEET_DADOS_URL nao configurada; usando backup JSON")
            return

        worksheets = self._get_worksheets(
            settings.google_sheet_dados_url,
            settings.google_sheet_dados_worksheet,
        )
        if not worksheets:
            logger.warning("Nao foi possivel acessar Google Sheets de dados; usando backup JSON")
            return

        expected_header = self._incoming_sheet_header()
        worksheet = worksheets[0]
        current_header = [value.strip() for value in worksheet.row_values(1)]
        if current_header != expected_header:
            logger.info("Atualizando colunas da planilha de interacoes para novo formato")
            self._replace_incoming_google_sheet(self._read_json(self.incoming_messages_file))
            return

        row = [
            entry["data_hora"],
            entry["student_name"],
            entry["class_name"],
            entry["ra"],
            entry["tipo_responsavel"],
            entry["numero_chamado"],
            entry["identificador_remetente"],
            entry["mensagem"],
            entry["intencao"],
            entry["motivo"],
            entry["observacao"],
            entry["campaign_id"],
            entry["origem"],
        ]

        try:
            worksheet.append_row(row)
            logger.info(
                "salvo no sheets | telefone=%s | class=%s",
                entry["telefone"],
                entry["intencao"] or entry["classificacao"],
            )
        except Exception as exc:
            logger.error("falha no sheets | usando backup JSON | erro=%s", exc)

    def resolver_nome_aluno(
        self,
        telefone: str,
        *,
        push_name: str = "",
    ) -> str:
        context = self.resolver_contexto_aluno(telefone, student_name="")
        return context.get("student_name", "")

    def resolver_contexto_aluno(
        self,
        telefone: str,
        *,
        student_name: str = "",
        push_name: str = "",
        data_hora: str = "",
    ) -> dict[str, str]:
        normalized_phone = self._normalize_phone_lookup(telefone)
        if normalized_phone:
            contact_context = self._find_contact_context_by_phone(normalized_phone)
            if contact_context:
                return self._merge_with_consolidated_context(contact_context)

            campaign_name = self._find_student_name_in_sent_campaigns(normalized_phone)
            if campaign_name:
                return self._merge_with_consolidated_context(
                    {
                        "student_name": campaign_name,
                        "class_name": "",
                        "ra": "",
                        "tipo_responsavel": "",
                        "numero_chamado": self._find_sent_phone_by_student_name(campaign_name),
                    }
                )

        recent_outbound_context = self._find_context_from_recent_outbound(
            push_name=push_name,
            data_hora=data_hora,
        )
        if recent_outbound_context:
            return recent_outbound_context

        if student_name.strip():
            return self._merge_with_consolidated_context(
                {
                    "student_name": student_name.strip(),
                    "class_name": "",
                    "ra": "",
                    "tipo_responsavel": "",
                    "numero_chamado": "",
                }
            )

        return {
            "student_name": "",
            "class_name": "",
            "ra": "",
            "tipo_responsavel": "",
            "numero_chamado": "",
        }

    def limpar_interacoes_salvas(self) -> int:
        entries = self._read_json(self.incoming_messages_file)
        cleaned_entries = self._clean_interaction_entries(entries)
        self._write_json(self.incoming_messages_file, cleaned_entries)
        self._replace_incoming_google_sheet(cleaned_entries)
        return len(cleaned_entries)

    def _clean_interaction_entries(self, entries: list[dict[str, Any]]) -> list[dict[str, Any]]:
        selected_entries: dict[str, dict[str, Any]] = {}
        anonymous_entries: list[dict[str, Any]] = []
        outbound_entries: list[dict[str, Any]] = []

        for entry in entries:
            raw_payload = entry.get("raw_payload") or {}
            if self._is_outbound_interaction(entry):
                outbound_entries.append(self._enrich_interaction_entry(entry))
                continue

            if not self._should_persist_incoming_payload(raw_payload):
                continue

            message_id = self._extract_message_id(raw_payload)
            if not message_id:
                anonymous_entries.append(self._enrich_interaction_entry(entry))
                continue

            current_best = selected_entries.get(message_id)
            if current_best is None or self._score_interaction_entry(entry) >= self._score_interaction_entry(
                current_best
            ):
                selected_entries[message_id] = entry

        cleaned = [self._enrich_interaction_entry(entry) for entry in selected_entries.values()]
        cleaned.extend(anonymous_entries)
        cleaned.extend(outbound_entries)
        cleaned.sort(key=lambda item: str(item.get("data_hora", "")))
        return cleaned

    def interacao_ja_registrada(self, payload: dict[str, Any]) -> bool:
        message_id = self._extract_message_id(payload)
        if not message_id:
            return False
        for entry in self._read_json(self.incoming_messages_file):
            existing_payload = entry.get("raw_payload") or {}
            if self._extract_message_id(existing_payload) == message_id:
                return True
        return False

    def _replace_incoming_google_sheet(self, entries: list[dict[str, Any]]) -> None:
        if not settings.google_sheet_dados_url:
            logger.warning("GOOGLE_SHEET_DADOS_URL nao configurada; limpeza apenas local")
            return

        worksheets = self._get_worksheets(
            settings.google_sheet_dados_url,
            settings.google_sheet_dados_worksheet,
        )
        if not worksheets:
            logger.warning("Nao foi possivel acessar Google Sheets de dados durante limpeza")
            return

        worksheet = worksheets[0]
        header = self._incoming_sheet_header()
        rows = [
            [
                entry.get("data_hora", ""),
                entry.get("student_name", ""),
                entry.get("class_name", ""),
                entry.get("ra", ""),
                entry.get("tipo_responsavel", ""),
                entry.get("numero_chamado", entry.get("telefone", "")),
                entry.get("identificador_remetente", entry.get("telefone", "")),
                entry.get("mensagem", ""),
                entry.get("intencao", ""),
                entry.get("motivo", ""),
                entry.get("observacao", ""),
                entry.get("campaign_id", ""),
                entry.get("origem", "whatsapp"),
            ]
            for entry in entries
        ]

        try:
            worksheet.clear()
            worksheet.append_row(header)
            if rows:
                worksheet.append_rows(rows)
            logger.info("Planilha de interacoes limpa e regravada com %s registros", len(rows))
        except Exception as exc:
            logger.error("Falha ao regravar planilha de interacoes: %s", exc)

    def _enrich_interaction_entry(self, entry: dict[str, Any]) -> dict[str, Any]:
        phone = str(entry.get("telefone", "")).strip()
        cleaned_entry = dict(entry)
        context = self.resolver_contexto_aluno(
            phone,
            student_name=str(entry.get("student_name", "")).strip(),
            push_name=str((entry.get("raw_payload") or {}).get("data", {}).get("pushName", "")).strip(),
            data_hora=str(entry.get("data_hora", "")).strip(),
        )
        fallback_numero_chamado = str(entry.get("numero_chamado", "")).strip() or str(
            entry.get("telefone", "")
        ).strip()
        cleaned_entry["student_name"] = str(entry.get("student_name", "")).strip() or context.get(
            "student_name",
            "",
        )
        cleaned_entry["class_name"] = str(entry.get("class_name", "")).strip() or context.get(
            "class_name",
            "",
        )
        cleaned_entry["ra"] = str(entry.get("ra", "")).strip() or context.get("ra", "")
        cleaned_entry["tipo_responsavel"] = str(entry.get("tipo_responsavel", "")).strip() or context.get(
            "tipo_responsavel",
            "",
        )
        cleaned_entry["numero_chamado"] = context.get(
            "numero_chamado", ""
        ) or fallback_numero_chamado
        cleaned_entry["identificador_remetente"] = str(
            entry.get("identificador_remetente", "")
        ).strip() or str(entry.get("telefone", "")).strip()
        return cleaned_entry

    @staticmethod
    def _score_interaction_entry(entry: dict[str, Any]) -> tuple[int, int, str]:
        raw_payload = entry.get("raw_payload") or {}
        data = raw_payload.get("data", {}) if isinstance(raw_payload, dict) else {}
        status = str(data.get("status", "")).strip().upper()
        push_name = str(data.get("pushName", "")).strip()
        status_score = 1 if status and status != "ERROR" else 0
        push_name_score = 1 if push_name else 0
        return (status_score, push_name_score, str(entry.get("data_hora", "")))

    def _find_student_name_in_sent_campaigns(self, telefone: str) -> str:
        campaigns_dir = self.base_path / "campaigns"
        if not campaigns_dir.exists():
            return ""

        for file_path in sorted(campaigns_dir.glob("*_sent.json"), reverse=True):
            for item in self._read_json(file_path):
                item_phone = self._normalize_phone_lookup(item.get("phone"))
                if item_phone == telefone:
                    return str(item.get("student_name", "")).strip()
        return ""

    def _find_sent_phone_by_student_name(self, student_name: str) -> str:
        campaigns_dir = self.base_path / "campaigns"
        if not campaigns_dir.exists():
            return ""

        normalized_student_name = self._normalize_text_lookup(student_name)
        for file_path in sorted(campaigns_dir.glob("*_sent.json"), reverse=True):
            for item in self._read_json(file_path):
                if self._normalize_text_lookup(item.get("student_name")) != normalized_student_name:
                    continue
                return self._strip_whatsapp_suffix(str(item.get("phone", "")).strip())
        return ""

    def _find_student_name_in_contacts_by_phone(self, telefone: str) -> str:
        for contact in self.carregar_contatos():
            phones = [
                self._normalize_phone_lookup(contact.get("phone1")),
                self._normalize_phone_lookup(contact.get("phone2")),
                self._normalize_phone_lookup(contact.get("phone3")),
            ]
            if telefone in phones:
                return str(contact.get("student_name", "")).strip()
        return ""

    def _find_contact_context_by_phone(self, telefone: str) -> dict[str, str]:
        for contact in self.carregar_contatos():
            for index in range(1, 4):
                if self._normalize_phone_lookup(contact.get(f"phone{index}")) != telefone:
                    continue
                return {
                    "student_name": str(contact.get("student_name", "")).strip(),
                    "class_name": str(contact.get("class_name", "")).strip(),
                    "ra": str(contact.get("ra", "")).strip(),
                    "tipo_responsavel": str(contact.get(f"responsible_type{index}", "")).strip(),
                    "numero_chamado": str(contact.get(f"phone{index}", "")).strip(),
                }
        return {}

    def _merge_with_consolidated_context(self, context: dict[str, str]) -> dict[str, str]:
        merged = {
            "student_name": str(context.get("student_name", "")).strip(),
            "class_name": str(context.get("class_name", "")).strip(),
            "ra": str(context.get("ra", "")).strip(),
            "tipo_responsavel": str(context.get("tipo_responsavel", "")).strip(),
            "numero_chamado": str(context.get("numero_chamado", "")).strip(),
        }
        consolidated = self._find_student_context_in_consolidated(
            merged["student_name"],
            merged["ra"],
        )
        if consolidated:
            merged["student_name"] = merged["student_name"] or consolidated.get("student_name", "")
            merged["class_name"] = merged["class_name"] or consolidated.get("class_name", "")
            merged["ra"] = merged["ra"] or consolidated.get("ra", "")
        return merged

    def _find_student_context_in_consolidated(
        self,
        student_name: str,
        ra: str,
    ) -> dict[str, str]:
        report_path = Path(settings.consolidated_report_path)
        if not report_path.exists():
            return {}

        consolidated_index = self._load_consolidated_index(report_path)
        if not consolidated_index:
            return {}

        normalized_ra = self._normalize_ra_lookup(ra)
        if normalized_ra and normalized_ra in consolidated_index:
            return dict(consolidated_index[normalized_ra])

        normalized_name = self._normalize_text_lookup(student_name)
        if normalized_name and normalized_name in consolidated_index:
            return dict(consolidated_index[normalized_name])

        return {}

    def _load_consolidated_index(self, report_path: Path) -> dict[str, dict[str, str]]:
        cache_path = str(report_path.resolve())
        if self._consolidated_cache is not None and self._consolidated_cache_path == cache_path:
            return self._consolidated_cache

        try:
            workbook = load_workbook(report_path, data_only=True)
        except Exception as exc:
            logger.warning("Falha ao abrir consolidado para enriquecer interacao: %s", exc)
            return {}

        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        header_index = self._find_consolidated_header_index(rows)
        if header_index < 0:
            return {}

        headers = [self._normalize_column_name(value) for value in rows[header_index]]
        name_index = self._find_column_index(headers, {"nome"})
        ra_index = self._find_column_index(headers, {"ra"})
        class_index = self._find_column_index(headers, {"turma"})
        if name_index < 0:
            return {}

        index: dict[str, dict[str, str]] = {}
        for row in rows[header_index + 1 :]:
            row_name = str(self._get_row_value(row, name_index)).strip()
            row_ra = str(self._get_row_value(row, ra_index)).strip() if ra_index >= 0 else ""
            row_class = str(self._get_row_value(row, class_index)).strip() if class_index >= 0 else ""
            if not row_name:
                continue

            context = {
                "student_name": row_name,
                "class_name": row_class,
                "ra": self._format_ra_from_consolidated(row_ra),
            }
            normalized_name = self._normalize_text_lookup(row_name)
            normalized_ra = self._normalize_ra_lookup(row_ra)
            if normalized_name:
                index[normalized_name] = context
            if normalized_ra:
                index[normalized_ra] = context

        self._consolidated_cache = index
        self._consolidated_cache_path = cache_path
        return index

    def _find_context_from_recent_outbound(
        self,
        *,
        push_name: str,
        data_hora: str,
    ) -> dict[str, str]:
        normalized_push_name = self._normalize_text_lookup(push_name)
        entries = self._read_json(self.incoming_messages_file)
        inbound_time = self._parse_iso_datetime(data_hora)
        candidates: list[dict[str, str]] = []

        for entry in reversed(entries):
            if str(entry.get("origem", "")).strip().lower() != "whatsapp_outbound":
                continue

            student_name = str(entry.get("student_name", "")).strip()
            if not student_name:
                continue

            entry_time = self._parse_iso_datetime(str(entry.get("data_hora", "")).strip())
            if inbound_time and entry_time:
                if entry_time > inbound_time:
                    continue
                if inbound_time - entry_time > timedelta(hours=12):
                    continue

            normalized_student = self._normalize_text_lookup(student_name)
            if normalized_push_name and self._push_name_matches_student(
                normalized_push_name,
                normalized_student,
            ):
                return self._merge_with_consolidated_context(
                    {
                        "student_name": student_name,
                        "class_name": str(entry.get("class_name", "")).strip(),
                        "ra": str(entry.get("ra", "")).strip(),
                        "tipo_responsavel": str(entry.get("tipo_responsavel", "")).strip(),
                        "numero_chamado": str(
                            entry.get("numero_chamado", entry.get("telefone", ""))
                        ).strip(),
                    }
                )

            candidates.append(
                {
                    "student_name": student_name,
                    "class_name": str(entry.get("class_name", "")).strip(),
                    "ra": str(entry.get("ra", "")).strip(),
                    "tipo_responsavel": str(entry.get("tipo_responsavel", "")).strip(),
                    "numero_chamado": str(
                        entry.get("numero_chamado", entry.get("telefone", ""))
                    ).strip(),
                }
            )
            if len(candidates) >= 5:
                break

        if len(candidates) == 1:
            return self._merge_with_consolidated_context(candidates[0])

        return {}

    @staticmethod
    def _should_persist_incoming_payload(payload: dict[str, Any]) -> bool:
        event = str(payload.get("event", "")).strip().lower()
        data = payload.get("data", {})
        key = data.get("key", {}) if isinstance(data, dict) else {}
        if event != "messages.upsert":
            return False
        if bool(key.get("fromMe")):
            return False
        return True

    @staticmethod
    def _extract_message_id(payload: dict[str, Any]) -> str:
        data = payload.get("data", {})
        key = data.get("key", {}) if isinstance(data, dict) else {}
        return str(key.get("id", "")).strip()

    @staticmethod
    def _normalize_phone_lookup(value: Any) -> str:
        text = str(value or "").replace("@s.whatsapp.net", "").replace("+", "").strip()
        digits = "".join(char for char in text if char.isdigit())
        if len(digits) > 11 and digits.startswith("55"):
            digits = digits[2:]
        return digits if len(digits) >= 10 else ""

    @staticmethod
    def _normalize_text_lookup(value: Any) -> str:
        text = str(value or "").strip().lower()
        text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
        return " ".join(text.split())

    @staticmethod
    def _normalize_ra_lookup(value: Any) -> str:
        return "".join(char for char in str(value or "").upper() if char.isalnum())

    @staticmethod
    def _format_ra_from_consolidated(value: Any) -> str:
        return " ".join(str(value or "").strip().split())

    @staticmethod
    def _strip_whatsapp_suffix(value: str) -> str:
        return value.replace("@s.whatsapp.net", "").replace("+", "").strip()

    @staticmethod
    def _parse_iso_datetime(value: str) -> datetime | None:
        if not value:
            return None
        try:
            return datetime.fromisoformat(value)
        except ValueError:
            return None

    @staticmethod
    def _push_name_matches_student(push_name: str, student_name: str) -> bool:
        push_tokens = [token for token in push_name.split() if token]
        student_tokens = [token for token in student_name.split() if token]
        if not push_tokens or not student_tokens:
            return False
        return push_tokens[0] == student_tokens[0]

    @staticmethod
    def _is_outbound_interaction(entry: dict[str, Any]) -> bool:
        origem = str(entry.get("origem", "")).strip().lower()
        if origem == "whatsapp_outbound":
            return True

        raw_payload = entry.get("raw_payload") or {}
        if not isinstance(raw_payload, dict):
            return False
        event = str(raw_payload.get("event", "")).strip().lower()
        return event == "send.message"

    @staticmethod
    def _incoming_sheet_header() -> list[str]:
        return [
            "data_hora",
            "student_name",
            "class_name",
            "ra",
            "tipo_responsavel",
            "numero_chamado",
            "identificador_remetente",
            "mensagem",
            "intencao",
            "motivo",
            "observacao",
            "campaign_id",
            "origem",
        ]

    def _records_to_contacts(self, records: list[dict[str, Any]]) -> list[dict[str, str]]:
        contacts: list[dict[str, str]] = []
        for record in records:
            normalized_record = {
                self._normalize_column_name(key): value for key, value in record.items()
            }

            student_name = self._pick_first_value(
                normalized_record,
                ["student_name", "nome_do_aluno", "nome_aluno", "aluno", "nome"],
            )
            if not student_name:
                continue

            phone_columns = self._find_phone_columns(normalized_record)
            phone_values = [
                self._sanitize_phone(normalized_record.get(column))
                for column in phone_columns[:3]
            ]
            while len(phone_values) < 3:
                phone_values.append("")

            contacts.append(
                {
                    "student_name": student_name,
                    "class_name": self._pick_first_value(
                        normalized_record,
                        ["class_name", "turma", "classe"],
                    ),
                    "ra": self._build_ra_value(normalized_record),
                    "phone1": phone_values[0],
                    "phone2": phone_values[1],
                    "phone3": phone_values[2],
                    "responsible_type1": self._pick_first_value(
                        normalized_record,
                        ["responsavel_1", "responsavel1"],
                    ),
                    "responsible_type2": self._pick_first_value(
                        normalized_record,
                        ["responsavel_2", "responsavel2"],
                    ),
                    "responsible_type3": self._pick_first_value(
                        normalized_record,
                        ["responsavel_3", "responsavel3"],
                    ),
                }
            )

        return contacts

    def _find_phone_columns(self, record: dict[str, Any]) -> list[str]:
        preferred = [
            "phone1",
            "phone2",
            "phone3",
            "phone_1",
            "phone_2",
            "phone_3",
            "telefone1",
            "telefone2",
            "telefone3",
            "telefone_1",
            "telefone_2",
            "telefone_3",
            "celular1",
            "celular2",
            "celular3",
            "celular_1",
            "celular_2",
            "celular_3",
        ]
        selected = [column for column in preferred if column in record]
        if selected:
            return selected[:3]

        discovered = [
            column
            for column in record
            if any(
                token in column
                for token in ("telefone", "celular", "fone", "whatsapp", "phone")
            )
        ]
        return discovered[:3]

    @staticmethod
    def _pick_first_value(record: dict[str, Any], columns: list[str]) -> str:
        for column in columns:
            value = record.get(column)
            if value is None:
                continue
            text = str(value).strip()
            if text:
                return text
        return ""

    def _build_ra_value(self, record: dict[str, Any]) -> str:
        ra = self._pick_first_value(record, ["ra"])
        if not ra:
            return ""

        digits = "".join(char for char in str(ra) if char.isdigit())
        ra_base = digits.zfill(12) if digits else str(ra).strip().upper()
        dig = self._pick_first_value(record, ["dig_ra", "dig__ra", "digito_ra"])
        uf = self._pick_first_value(record, ["uf_ra"])

        formatted = ra_base
        if dig:
            formatted = f"{formatted}-{str(dig).strip().upper()}"
        if uf:
            formatted = f"{formatted}/{str(uf).strip().upper()}"
        return formatted

    @staticmethod
    def _sanitize_phone(value: Any) -> str:
        digits = "".join(char for char in str(value or "") if char.isdigit())
        if len(digits) < 10:
            return ""
        return digits

    @staticmethod
    def _normalize_column_name(value: Any) -> str:
        text = str(value or "").strip().lower()
        text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
        normalized = []
        for char in text:
            normalized.append(char if char.isalnum() else "_")
        return "".join(normalized).strip("_")

    @staticmethod
    def _find_consolidated_header_index(rows: list[tuple[Any, ...]]) -> int:
        for index, row in enumerate(rows):
            normalized = {
                LocalRepository._normalize_column_name(value)
                for value in row
                if str(value or "").strip()
            }
            if {"nome", "ra", "turma"}.issubset(normalized):
                return index
        return -1

    @staticmethod
    def _find_column_index(headers: list[str], options: set[str]) -> int:
        for index, header in enumerate(headers):
            if header in options:
                return index
        return -1

    @staticmethod
    def _get_row_value(row: tuple[Any, ...], index: int) -> Any:
        if index < 0 or index >= len(row):
            return ""
        return row[index]


repository = LocalRepository()
